from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.utils import get_column_letter


def is_orange(cell):
    """
    Check whether a cell has an orange background.
    """

    fill = cell.fill

    if fill.fill_type != "solid":
        return False

    color = fill.fgColor

    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:]

        try:
            r = int(rgb[0:2], 16)
            g = int(rgb[2:4], 16)
            b = int(rgb[4:6], 16)

            # Detect orange shades
            return (
                r >= 180
                and 70 <= g <= 220
                and b <= 150
                and r > g + 30
                and g > b
            )

        except ValueError:
            return False

    return False


def find_orange_row(ws):
    """
    Find the first row containing an orange cell.

    Returns:
        Row number if found
        None if not found
    """

    for row in ws.iter_rows():
        for cell in row:

            if is_orange(cell):
                return cell.row

    return None


def copy_cell(source_cell, target_cell):
    """
    Copy cell value and formatting.
    """

    target_cell.value = source_cell.value

    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def copy_row(source_ws, target_ws, source_row, target_row):
    """
    Copy one row from source worksheet.

    Original column A -> Output column B
    Original column B -> Output column C
    etc.

    Output column A is reserved for the sheet name/date.
    """

    for source_col in range(1, source_ws.max_column + 1):

        source_cell = source_ws.cell(
            row=source_row,
            column=source_col
        )

        # Shift one column to the right
        target_cell = target_ws.cell(
            row=target_row,
            column=source_col + 1
        )

        copy_cell(source_cell, target_cell)

    # Preserve row height
    if source_row in source_ws.row_dimensions:

        target_ws.row_dimensions[target_row].height = (
            source_ws.row_dimensions[source_row].height
        )


def process_excel(input_file, output_file):

    # ---------------------------------------------
    # OPEN SOURCE WORKBOOK
    # ---------------------------------------------

    source_wb = load_workbook(input_file)

    # ---------------------------------------------
    # CREATE OUTPUT WORKBOOK
    # ---------------------------------------------

    output_wb = Workbook()

    output_ws = output_wb.active
    output_ws.title = "Combined Data"

    output_row = 1

    # ---------------------------------------------
    # PROCESS EVERY SHEET
    # ---------------------------------------------

    for source_ws in source_wb.worksheets:

        print(f"\nProcessing sheet: {source_ws.title}")

        # -----------------------------------------
        # FIND ORANGE ROW
        # -----------------------------------------

        orange_row = find_orange_row(source_ws)

        if orange_row is None:

            print("  No orange cell found. Sheet skipped.")

            continue

        print(f"  Orange row found: {orange_row}")

        # -----------------------------------------
        # COPY FROM ROW 4
        # UNTIL ROW BEFORE ORANGE ROW
        # -----------------------------------------

        for source_row in range(4, orange_row):

            # -------------------------------------
            # COLUMN A = SHEET NAME / DATE
            # -------------------------------------

            date_cell = output_ws.cell(
                row=output_row,
                column=1
            )

            date_cell.value = source_ws.title

            # -------------------------------------
            # ORIGINAL DATA STARTS FROM COLUMN B
            # -------------------------------------

            copy_row(
                source_ws,
                output_ws,
                source_row,
                output_row
            )

            output_row += 1

        print(
            f"  Copied rows 4 to {orange_row - 1}"
        )

    # ---------------------------------------------
    # COPY COLUMN WIDTHS
    # ---------------------------------------------

    if source_wb.worksheets:

        first_ws = source_wb.worksheets[0]

        for col_index in range(
            1,
            first_ws.max_column + 1
        ):

            source_letter = get_column_letter(col_index)

            target_letter = get_column_letter(
                col_index + 1
            )

            width = first_ws.column_dimensions[
                source_letter
            ].width

            if width:
                output_ws.column_dimensions[
                    target_letter
                ].width = width

    # Width for Date column
    output_ws.column_dimensions["A"].width = 15

    # ---------------------------------------------
    # SAVE OUTPUT
    # ---------------------------------------------

    output_wb.save(output_file)

    print("\n--------------------------------")
    print("Processing completed successfully!")
    print(f"Output file: {output_file}")
    print("--------------------------------")


# =================================================
# CHANGE THESE TWO PATHS
# =================================================

input_file = r"C:\Users\YourName\Documents\DailyTasks.xlsx"

output_file = r"C:\Users\YourName\Documents\CombinedTasks.xlsx"


# =================================================
# RUN
# =================================================

process_excel(
    input_file,
    output_file
)
